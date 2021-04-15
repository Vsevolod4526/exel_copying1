from tkinter import *
from tkinter import messagebox
import os
import openpyxl
import xlrd
from openpyxl.styles import Font
from tkinter import filedialog as fd #askdirectory
programm_native_path=os.getcwd()
root=Tk()
root.geometry('600x400+600+300')
root.resizable(False, False)
root.title("Копирование данных из отчётов .xls")
b1=Button(text='Запуск копирования',width=30,height=5)
l1=Label(text='Укажите путь в папку с файлами .xls',font='Arial 16')
e1=Entry(width=70)
l2=Label(text='Укажите путь в папку для сохраняемого файла ',font='Arial 16')
e2=Entry(width=70)
l3=Label(text='                                    ',font='Arial 16')
b2Nova_directory=Button(text='Выбор папки с файлами .xls',width=30,height=2)
b3Nova_save_directory=Button(text='Выбор папки для сохранения файла',width=30,height=2)
l4=Label(text='Название для сохраняемого файла',font='Arial 16')
e3=Entry(width=40)

try:
    old_text_fail=open("directories.txt","r",encoding='utf-8')
    e1.insert(0,old_text_fail.readline()[:-1])
    e2.insert(0,old_text_fail.readline())
    old_text_fail.close()
except:
    print("Ошибка при нахождении txt файла с сохраненными директориями")




    
def vibor_fail_directory():
    file_directory = fd.askdirectory()
    e1.delete(0, last=END)
    e1.insert(0,file_directory)
def vibor_save_directory():
    save_file_directory=fd.askdirectory()
    e2.delete(0, last=END)
    e2.insert(0,save_file_directory)
    
 

    



def start_copying():
    b1['text']='Идёт копирование'
    b1['bg'] = '#00ff00'
    direct=e1.get()
    stock_direct=e2.get()
    
    list_of_files=[]
    list_of_xlsx_files=[]
    POZ_point=1
    try:
        os.chdir(direct)
    except OSError:
        print("Ошибка в имени директории не существует")
        messagebox.showerror("Ошибка","Ошибка в пути папки. Директории не существует")

    stock_EXEL_fail_workbook=openpyxl.Workbook()
    sheet_of_stock_EXEL = stock_EXEL_fail_workbook[stock_EXEL_fail_workbook.sheetnames[0]]
    row_sheet_stock_EXEL=2
    
    list_of_files=os.listdir(path=direct)
    for i in range(len(list_of_files)):
        if '.xls' in list_of_files[i]:
            list_of_xlsx_files.append(list_of_files[i])
    print(list_of_xlsx_files)
    for i in range(len(list_of_xlsx_files)):
        workbook=xlrd.open_workbook(list_of_xlsx_files[i])#workbook=openpyxl.load_workbook(list_of_xlsx_files[i])
        sheet=workbook.sheet_by_index(0)#sheet=workbook[workbook.sheetnames[0]]
        last_row=sheet.nrows ##
        print(last_row)
        data=sheet.cell_value(6,7)
        nomer=int(sheet.cell_value(5,7))
        for j in range(18,last_row):
            list_of_cells=[]
            if type(sheet.cell_value(j,1))==float:
                list_of_cells=sheet.row_values(j, start_colx=0, end_colx=10)
                print(list_of_cells)

                for h in range(1 ,11):
                    sheet_of_stock_EXEL.cell(row = row_sheet_stock_EXEL, column = h).value=list_of_cells[h-1]
                sheet_of_stock_EXEL.cell(row = row_sheet_stock_EXEL, column = 12).value=data
                sheet_of_stock_EXEL.cell(row = row_sheet_stock_EXEL, column = 11).value=nomer
                row_sheet_stock_EXEL+=1
    sheet_of_stock_EXEL.cell(row = 1, column = 1).value="Поз."
    sheet_of_stock_EXEL.cell(row = 1, column = 2).value="Кол-во (шт.)"
    sheet_of_stock_EXEL.cell(row = 1, column = 3).value="Артикул"
    sheet_of_stock_EXEL.cell(row = 1, column = 4).value="Наименование"
    sheet_of_stock_EXEL.cell(row = 1, column = 5).value="Страна изготовитель."
    sheet_of_stock_EXEL.cell(row = 1, column = 6).value="Вес брутто (кг)"
    sheet_of_stock_EXEL.cell(row = 1, column = 7).value="Вес нетто (кг)"
    sheet_of_stock_EXEL.cell(row = 1, column = 8).value="Цена за ед."
    sheet_of_stock_EXEL.cell(row = 1, column = 9).value="Цена брутто"
    sheet_of_stock_EXEL.cell(row = 1, column = 10).value="Всего"
    sheet_of_stock_EXEL.cell(row = 1, column = 11).value="Номер"
    sheet_of_stock_EXEL.cell(row = 1, column = 12).value="Дата"

    for p in range(1,13):
        sheet_of_stock_EXEL.cell(row = 1, column = p).font= Font(bold=True)
                
    try:
        os.chdir(stock_direct)
    except OSError:
        print("Ошибка в имени директории не существует")
        messagebox.showerror("Ошибка","Ошибка в пути папки для стокого файла. Директории не существует")
    stock_fail_name=e3.get()
    stock_EXEL_fail_workbook.save(stock_fail_name+'.xlsx')


    os.chdir(programm_native_path)
    write_txt_directories=open("directories.txt","w+",encoding='utf-8')
    write_txt_directories.write(e1.get())
    write_txt_directories.write("\n")
    write_txt_directories.write(e2.get())
    write_txt_directories.close()
    l3['text']='Копирование завершено'
    b1['text']='Копирование завершено'
    
b1.config(command=start_copying)
b2Nova_directory.config(command=vibor_fail_directory)
b3Nova_save_directory.config(command=vibor_save_directory)

l1.pack()
e1.pack()
b2Nova_directory.pack()
l2.pack()
e2.pack()
b3Nova_save_directory.pack()
l4.pack()
e3.pack()
b1.pack()
l3.pack()

root.mainloop()
